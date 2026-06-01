"""
Generate the spike fixture xlsx. Run once; commit the output.

  pip install openpyxl
  python spike/fixtures/build_fixture.py

The output is committed at spike/fixtures/test.xlsx.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Hoja1"

# Header row with formatting (R5: color).
header = ["Producto", "Cantidad", "Precio", "Total"]
for col, value in enumerate(header, start=1):
    cell = ws.cell(row=1, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor="FFD966")
    cell.font = Font(bold=True)

# Merged title row above header (R5: merged cells).
ws.insert_rows(1)
ws.cell(row=1, column=1, value="Reporte Q3 2026").font = Font(size=14, bold=True)
ws.merge_cells("A1:D1")

# 1000 data rows (R2.1 fuel — but here we want >= 1000 *cells*, not rows).
import random
random.seed(42)
for i in range(3, 253):  # 250 rows × 4 cols = 1000 cells
    ws.cell(row=i, column=1, value=f"SKU-{i-2:04d}")
    qty = random.randint(1, 20)
    price = round(random.uniform(5, 200), 2)
    ws.cell(row=i, column=2, value=qty)
    ws.cell(row=i, column=3, value=price)
    # Formula cell (R5: formula).
    ws.cell(row=i, column=4, value=f"=B{i}*C{i}")

# Grand total formula at the bottom.
total_row = 253
ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
ws.cell(row=total_row, column=4, value="=SUM(D3:D252)").font = Font(bold=True)

out = Path(__file__).parent / "test.xlsx"
wb.save(out)
print(f"wrote {out} ({out.stat().st_size} bytes)")
